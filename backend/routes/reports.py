import io
from flask import Blueprint, jsonify, send_file
from database.db import SessionLocal
from database.models import AttendanceSession, Attendance, Student

reports_bp = Blueprint("reports", __name__)


@reports_bp.route("/sessions", methods=["GET"])
def get_sessions():
    """List all attendance sessions with present / absent counts."""
    db = SessionLocal()
    try:
        sessions = db.query(AttendanceSession).order_by(AttendanceSession.id.desc()).all()
        total_students = db.query(Student).count()

        result = []
        for s in sessions:
            present = db.query(Attendance).filter_by(session_id=s.id).count()
            absent = max(total_students - present, 0)
            result.append({
                "session_id": s.id,
                "date": s.date,
                "start_time": s.start_time.strftime("%H:%M:%S") if s.start_time else None,
                "end_time": s.end_time.strftime("%H:%M:%S") if s.end_time else None,
                "status": s.status,
                "present": present,
                "absent": absent,
                "total": total_students,
                "rate": round((present / total_students * 100), 1) if total_students > 0 else 0
            })

        return jsonify(result)
    finally:
        db.close()


@reports_bp.route("/session/<int:session_id>", methods=["GET"])
def get_session_detail(session_id):
    """Full attendance sheet for a session: every student with Present / Absent status."""
    db = SessionLocal()
    try:
        session = db.query(AttendanceSession).filter_by(id=session_id).first()
        if not session:
            return jsonify({"error": "Session not found"}), 404

        all_students = db.query(Student).all()

        # Set of student ids that were marked present in this session
        present_ids = {
            row.student_id
            for row in db.query(Attendance).filter_by(session_id=session_id).all()
        }

        sheet = []
        for student in all_students:
            marked = db.query(Attendance).filter_by(
                session_id=session_id,
                student_id=student.id
            ).first()

            sheet.append({
                "student_id": student.id,
                "name": student.name,
                "roll_number": student.roll_number,
                "department": student.department,
                "year": student.year,
                "status": "Present" if student.id in present_ids else "Absent",
                "time": marked.time if marked else None,
            })

        return jsonify({
            "session_id": session.id,
            "date": session.date,
            "status": session.status,
            "students": sheet
        })
    finally:
        db.close()


def generate_excel_for_session(session_id):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return None

    db = SessionLocal()
    try:
        session = db.query(AttendanceSession).filter_by(id=session_id).first()
        if not session:
            return None

        all_students = db.query(Student).all()

        present_map = {
            row.student_id: row.time
            for row in db.query(Attendance).filter_by(session_id=session_id).all()
        }

        # ── Build workbook in memory ──────────────────────────────────────────
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Session {session_id}"

        # Header row styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F46E5")  # indigo
        center = Alignment(horizontal="center")

        headers = ["#", "Roll Number", "Student Name", "Department", "Year", "Status", "Time"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        # Data rows
        for i, student in enumerate(all_students, start=1):
            is_present = student.id in present_map
            status = "Present" if is_present else "Absent"
            time_val = present_map.get(student.id, "—")
            ws.append([i, student.roll_number, student.name,
                        student.department, str(student.year) or "—", status, time_val])

            # Colour-code status cell
            status_cell = ws.cell(row=i + 1, column=6)
            status_cell.font = Font(bold=True,
                                     color="10B981" if is_present else "EF4444")

        # Auto-fit column widths
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

        # ── Stream to Excel Bytes ─────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
    finally:
        db.close()

@reports_bp.route("/export/<int:session_id>", methods=["GET"])
def export_session(session_id):
    """Download attendance sheet as .xlsx — all data in-memory (Render safe, no disk writes)."""
    db = SessionLocal()
    try:
        session = db.query(AttendanceSession).filter_by(id=session_id).first()
        if not session:
            return jsonify({"error": "Session not found"}), 404

        buf = generate_excel_for_session(session_id)
        if not buf:
            return jsonify({"error": "openpyxl not installed or session missing"}), 500

        filename = f"attendance_session_{session_id}_{session.date}.xlsx"
        return send_file(
            buf,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    finally:
        db.close()
